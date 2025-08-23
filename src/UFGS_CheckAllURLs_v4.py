# Auto-generated script from UFGS_CheckAllURLs_v4.ipynb
# In[1]:
# Main Script Logic for Google Colab - Concurrency Comparison Test

# Pip Installs (uncomment and run this line in a cell if needed)
# !pip install reachable==0.7.0 httpx fake-useragent lxml pandas openpyxl

import zipfile
import os
import re # For regular expressions
import pandas as pd
import httpx
import time
import shutil
import sys
import urllib.parse
import asyncio
import traceback # For detailed error printing

# --- Reachable Library Imports ---
async_components_imported_successfully = False
is_reachable_async_func = None
AsyncClient_class = None
TaskPool_class = None

try:
    from reachable import is_reachable_async as imported_is_reachable_async
    is_reachable_async_func = imported_is_reachable_async
    from reachable.client import AsyncClient as imported_AsyncClient
    AsyncClient_class = imported_AsyncClient
    from reachable.pool import TaskPool as imported_TaskPool
    TaskPool_class = imported_TaskPool
    async_components_imported_successfully = True
    print("Successfully imported async components from 'reachable'.")
except ImportError as e:
    print(f"CRITICAL ERROR: Could not import required async components from 'reachable': {e}")
    print("Please ensure 'reachable==0.7.0' (AlexMili/Reachable on PyPI) is installed correctly.")
    is_reachable_async_func = lambda *args, **kwargs: asyncio.sleep(0)
    class AsyncClientPlaceholder:
        async def __aenter__(self):
            await asyncio.sleep(0)
            return self

        async def __aexit__(self, *args):
            await asyncio.sleep(0)

    AsyncClient_class = AsyncClientPlaceholder

    class TaskPoolPlaceholder:
        def __init__(self, *args, **kwargs):
            self._results = []

        async def put(self, coro):
            await asyncio.sleep(0)
            self._results.append(None)

        async def join(self):
            await asyncio.sleep(0)

    TaskPool_class = TaskPoolPlaceholder

# Imports for Excel formatting
from openpyxl.styles import Font, PatternFill, colors, Alignment, Border, Side, NumberFormatDescriptor
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- Configuration ---
DO_FULL_PROCESSING = True
debug = False # Set to True for verbose output, False for cleaner comparison output

# --- Constants ---
ufgs_download_url = "https://www.wbdg.org/FFC/DOD/UFGS/UFGS_M.zip"
zip_file_path = '/content/UFGS.zip'
extract_path = '/content/sec_files'
output_excel_path_base = '/content/SEC_URL_references_with_certainty' # Base name for output files

TARGET_SPECS_DATA = [
    ("02 84 16", "Electrical"), ("02 84 33", "Electrical"), ("08 34 49.00 20", "Electrical"),
    ("08 71 63.10", "Electrical"), ("26 05 13.00 10", "Electrical"), ("26 05 19.00 10", "Electrical"),
    ("26 05 33", "Electrical"), ("26 05 48", "Electrical"), ("26 05 73", "Electrical"),
    ("26 08 00", "Electrical"), ("26 11 13.00 20", "Electrical"), ("26 11 14.00 10", "Electrical"),
    ("26 11 16", "Electrical"), ("26 12 19", "Electrical"), ("26 12 21", "Electrical"),
    ("26 13 00", "Electrical"), ("26 13 01", "Electrical"), ("26 13 02", "Electrical"),
    ("26 13 13", "Electrical"), ("26 13 14", "Electrical"), ("26 13 32", "Electrical"),
    ("26 19 13", "Electrical"), ("26 20 00", "Electrical"), ("26 22 00.00 10", "Electrical"),
    ("26 23 00", "Electrical"), ("26 24 13", "Electrical"), ("26 27 29", "Electrical"),
    ("26 28 00.00 10", "Electrical"), ("26 29 01.00 10", "Electrical"), ("26 29 02.00 10", "Electrical"),
    ("26 29 23", "Electrical"), ("26 31 00", "Electrical"), ("26 32 15", "Electrical"),
    ("26 33 00", "Electrical"), ("26 33 53", "Electrical"), ("26 35 43", "Electrical"),
    ("26 35 44", "Electrical"), ("26 36 23", "Electrical"), ("26 41 00", "Electrical"),
    ("26 42 13", "Electrical"), ("26 42 15", "Electrical"), ("26 42 17", "Electrical"),
    ("26 42 19.00 10", "Electrical"), ("26 51 00", "Electrical"), ("26 55 53", "Electrical"),
    ("26 56 00", "Electrical"), ("26 56 20", "Electrical"), ("27 05 13.43", "Electrical"),
    ("27 05 26", "Electrical"), ("27 05 29.00 10", "Electrical"), ("27 10 00", "Electrical"),
    ("27 41 00", "Electrical"), ("27 51 16", "Electrical"), ("27 51 23", "Electrical"),
    ("27 53 19", "Electrical"), ("28 08 10", "Electrical"), ("28 10 05", "Electrical"),
    ("28 20 02", "Electrical"), ("33 71 01", "Electrical"), ("33 71 02", "Electrical"),
    ("33 82 00", "Electrical"), ("34 60 13", "Electrical"), ("35 20 20", "Electrical"),
    ("48 14 00", "Electrical"), ("48 15 00", "Electrical"), ("48 16 00", "Electrical")
]

url_pattern = re.compile(r'<URL(?:\s+HREF="([^"]*)")?[^>]*>(.*?)</URL>', re.IGNORECASE | re.DOTALL)
scn_pattern = re.compile(r'<SCN>(.*?)</SCN>', re.IGNORECASE | re.DOTALL)
stl_pattern = re.compile(r'<STL>(.*?)</STL>', re.IGNORECASE | re.DOTALL)
dte_pattern = re.compile(r'<DTE>(.*?)</DTE>', re.IGNORECASE | re.DOTALL)
pra_pattern = re.compile(r'<PRA>(.*?)</PRA>', re.IGNORECASE | re.DOTALL)

# --- Helper Functions ---
def debug_print(message):
    if debug:
        print(message)

def is_email(url_to_check):
    if not isinstance(url_to_check, str):
        return False
    return url_to_check.startswith("mailto:") or '@' in url_to_check

def is_wbdg_url(url_to_check):
    if isinstance(url_to_check, str):
        return "wbdg.org" in url_to_check.lower()
    return False

def parse_reachable_result(result_dict: dict, original_url_input: str, current_run_error_counts: dict):
    validation_status = "FAIL"
    current_status_code = None
    current_final_url = original_url_input
    error_message = "Initial parsing error"
    check_certainty = "HIGH"

    if not isinstance(result_dict, dict):
        error_message = f"Invalid result: expected dict, got {type(result_dict)}"
        if original_url_input not in current_run_error_counts:
            current_run_error_counts[original_url_input] = 0
        current_run_error_counts[original_url_input] += 1
        return ("PROCESSING_ERROR", None, original_url_input, error_message, is_wbdg_url(original_url_input), "LOW")

    url_str_from_dict = result_dict.get('original_url', original_url_input)
    wbdg_check = is_wbdg_url(url_str_from_dict)

    if result_dict.get('is_email_flag_for_parser'):
         return ("EMAIL", None, url_str_from_dict, None, wbdg_check, "HIGH")
    if result_dict.get('is_invalid_flag_for_parser'):
        return ("INVALID", None, url_str_from_dict, "Invalid URL format (pre-checked)", wbdg_check, "HIGH")
    if result_dict.get('error_detail_worker') or result_dict.get('error_detail_setup'):
        error_message = f"Worker/Setup Error: {result_dict.get('error_name', 'Unknown')} - {result_dict.get('error_detail_worker') or result_dict.get('error_detail_setup')}"
        validation_status = result_dict.get('custom_status', "FAIL_WORKER")
        current_final_url = url_str_from_dict
        current_status_code = result_dict.get('status_code')
        if original_url_input not in current_run_error_counts:
            current_run_error_counts[original_url_input] = 0
        current_run_error_counts[original_url_input] += 1
        return (validation_status, current_status_code, current_final_url, error_message, wbdg_check, "HIGH")

    success = result_dict.get('success', False)
    debug_print(f"    Parsing result for '{url_str_from_dict}'. Library success: {success}, Error Name: {result_dict.get('error_name')}, Status Code: {result_dict.get('status_code')}")
    status_code_from_dict = result_dict.get('status_code')
    error_name_from_dict = result_dict.get('error_name')
    current_final_url = result_dict.get('final_url', url_str_from_dict)
    current_status_code = status_code_from_dict if status_code_from_dict != -1 else None

    response_text_content = None
    httpx_response_obj = result_dict.get('response')
    if httpx_response_obj and hasattr(httpx_response_obj, 'text'):
        try:
            response_text_content = httpx_response_obj.text
            if response_text_content is None: response_text_content = ""
        except Exception as e_text:
            debug_print(f"    Error accessing .text from response object for {url_str_from_dict}: {e_text}")
            response_text_content = ""
        if debug and response_text_content:
            debug_print(f"    Response text for {url_str_from_dict} (first 100): '{response_text_content[:100]}'")
    elif debug and result_dict.get('include_response'):
         debug_print(f"    include_response=True but 'response' key was '{httpx_response_obj}' (type: {type(httpx_response_obj)}) and unusable for text for {url_str_from_dict}. Keys in dict: {list(result_dict.keys())}")

    is_parked = result_dict.get('is_parking_domain', False)
    parked_domain_message = "Domain may be parked." if is_parked else ""

    if wbdg_check and response_text_content:
        debug_print(f"    WBDG check for {url_str_from_dict} with text content.")
        text_content_lower = response_text_content.lower()
        wbdg_error_phrases = ["oops", "an error has occurred", "page can’t be found", "page not found", "error processing your request", "server error", "unable to process", "temporarily unavailable"]
        if any(phrase in text_content_lower for phrase in wbdg_error_phrases):
            validation_status = "WARN_WBDG_CONTENT_ERROR"
            error_message = "WBDG: Content check triggered an error page."
            if is_parked: error_message += f" {parked_domain_message}"
            check_certainty = "HIGH"
            if original_url_input not in current_run_error_counts:
                current_run_error_counts[original_url_input] = 0
            current_run_error_counts[original_url_input] += 1
            debug_print(f"    ==> Final parse_reachable_result for {original_url_input}: Status={validation_status}, Code={current_status_code}, FinalURL='{current_final_url}', Msg='{error_message}', WBDG={wbdg_check}, Certainty={check_certainty}")
            return (validation_status, current_status_code, current_final_url, error_message, wbdg_check, check_certainty)

    if success:
        validation_status = "PASS"
        error_message = parked_domain_message if is_parked else None
        check_certainty = "MEDIUM" if is_parked and validation_status == "PASS" else "HIGH"
    else:
        validation_status = "FAIL"
        inline_tag_detected_in_original = False
        if re.search(r'<[^>]+>', original_url_input):
            inline_tag_detected_in_original = True
            debug_print(f"    Inline tag detected in original URL: {original_url_input}")

        if error_name_from_dict == 'InvalidURL' and inline_tag_detected_in_original:
            error_message = "Failed. Library reported 'InvalidURL', likely due to inline tag(s) (e.g., <BRK/>) in the URL."
        elif error_name_from_dict:
            error_message = f"Failed. Error: {error_name_from_dict}."
            if inline_tag_detected_in_original:
                error_message += " Suspected inline tag(s) in URL may have contributed."
        else:
            error_message = "Failed (success=False)."
            if inline_tag_detected_in_original:
                error_message += " Suspected inline tag(s) in URL."

        if status_code_from_dict is not None:
            error_message += f" (Raw Status: {status_code_from_dict})"

        if is_parked:
            parked_note = parked_domain_message
            if error_message and not error_message.endswith('.'): error_message += "."
            error_message = f"{error_message} {parked_note}" if error_message else parked_note

        check_certainty = "HIGH"
        if original_url_input not in current_run_error_counts:
            current_run_error_counts[original_url_input] = 0
        current_run_error_counts[original_url_input] += 1

    debug_print(f"    ==> Final parse_reachable_result for {original_url_input}: Status={validation_status}, Code={current_status_code}, FinalURL='{current_final_url}', Msg='{error_message}', WBDG={wbdg_check}, Certainty={check_certainty}")
    return (validation_status, current_status_code, current_final_url, error_message, wbdg_check, check_certainty)

async def async_url_worker(url: str, client: AsyncClient_class, worker_params: dict):
    if not is_reachable_async_func or not callable(is_reachable_async_func):
        return {'original_url': url, 'success': False, 'error_name': 'LibFuncMissingAsync'}
    try:
        # Removed 'attempts' from direct pass as it caused TypeError.
        # Relying on library's default for attempts.
        params_for_call = {k: v for k, v in worker_params.items() if k != 'attempts'}
        return await is_reachable_async_func(url, client=client, **params_for_call)
    except TypeError as te:
        debug_print(f"  TypeError in async_url_worker for {url} calling is_reachable_async: {te}")
        return {'original_url': url, 'success': False, 'status_code': None, 'error_name': 'TypeErrorKwargs', 'error_detail_worker': str(te)}
    except Exception as e:
        debug_print(f"  Exception in async_url_worker for {url}: {type(e).__name__} - {str(e)}")
        return {'original_url': url, 'success': False, 'status_code': None, 'error_name': type(e).__name__, 'error_detail_worker': str(e)}

async def run_async_validators(urls_to_check: list, pool_size: int = 50):
    results = []
    if not async_components_imported_successfully or not AsyncClient_class or not TaskPool_class or AsyncClient_class.__name__ == 'AsyncClientPlaceholder' or TaskPool_class.__name__ == 'TaskPoolPlaceholder':
        print("ERROR: Async components not properly available. Cannot run async validation.")
        for url in urls_to_check: results.append({'original_url': url, 'success': False, 'error_name': 'AsyncSetupErrorMain'})
        return results

    debug_print(f"  run_async_validators: Initializing AsyncClient with library defaults (no explicit timeout argument).")
    async with AsyncClient_class() as client: # NO TIMEOUT ARGUMENT HERE
        tasks_pool = TaskPool_class(workers=pool_size)
        debug_print(f"  Submitting {len(urls_to_check)} URLs to TaskPool with {pool_size} workers...")

        worker_call_params = {
            "headers": None,
            "include_response": True,
            "check_parking_domain": True
            # "attempts" is removed as it caused TypeError with is_reachable_async for this library version
        }
        for url_item in urls_to_check: await tasks_pool.put(async_url_worker(url_item, client, worker_call_params))

        debug_print(f"  All tasks submitted. Awaiting TaskPool.join()...")
        await tasks_pool.join()
        debug_print(f"  TaskPool.join() completed.")

        if hasattr(tasks_pool, '_results'): results = tasks_pool._results
        elif hasattr(tasks_pool, 'results'): results = tasks_pool.results
        else:
            print("ERROR: Could not retrieve results from TaskPool.")
            results = [{'original_url': u, 'success': False, 'error_name': 'TaskPoolResultErrorMain'} for u in urls_to_check]
        debug_print(f"  Retrieved {len(results)} results from TaskPool.")
    return results

def parse_sec_file_for_urls(file_path, current_id_counter):
    debug_print(f"  Parsing for URLs: {os.path.basename(file_path)}")
    extracted_rows = []
    urls_found_in_file = 0
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file: content = file.read()
        for match in url_pattern.finditer(content):
            url_href, url_txt = (match.group(1) if match.group(1) else None), (match.group(2).strip() if match.group(2) else '')
            primary_url = None
            if url_href: primary_url = url_href.strip()
            elif url_txt and '.' in url_txt:
                cleaned_url_txt = re.sub(r'<[^>]+>', '', url_txt.strip()).strip()
                if ' ' not in cleaned_url_txt and (cleaned_url_txt.count('.') >= 1 or '@' in cleaned_url_txt or cleaned_url_txt.lower().startswith("www.")): primary_url = cleaned_url_txt
                elif cleaned_url_txt.lower().startswith("http"): primary_url = cleaned_url_txt
            if not primary_url:
                continue
            char_offset, line_number = match.start(), content.count('\n', 0, match.start()) + 1
            extracted_rows.append((current_id_counter, os.path.basename(file_path), line_number, char_offset, url_txt, url_href, primary_url))
            current_id_counter += 1
            urls_found_in_file += 1
    except Exception as e:
        debug_print(f"  Error parsing URLs in {os.path.basename(file_path)}: {e}")
    debug_print(f"  Finished parsing URLs for {os.path.basename(file_path)}. Found {urls_found_in_file} URLs.")
    return extracted_rows, current_id_counter

def extract_file_metadata(file_path):
    filename = os.path.basename(file_path)
    metadata = {'Filename': filename, 'Section Number': "Not Found", 'Section Title': "Not Found", 'Date': "Not Found", 'Preparing Authority': "Not Found" }
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file: content_start = file.read(8192)
        if (m := scn_pattern.search(content_start)): metadata['Section Number'] = m.group(1).strip()
        if (m := stl_pattern.search(content_start)): metadata['Section Title'] = ' '.join(m.group(1).split()).strip()
        if (m := dte_pattern.search(content_start)): metadata['Date'] = m.group(1).strip()
        if (m := pra_pattern.search(content_start)): metadata['Preparing Authority'] = ' '.join(m.group(1).split()).strip()
    except Exception as e:
        debug_print(f"  Error extracting metadata from {filename}: {e}")
    return metadata

def auto_fit_columns(worksheet: Worksheet, columns_to_fit=None, max_width=100, padding=3):
    col_indices = [ord(c.upper()) - ord('A') + 1 for c in columns_to_fit] if columns_to_fit else range(1, worksheet.max_column + 1)
    for col_idx in col_indices:
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            is_merged, cell_val_len = False, 0
            for merged_range_obj in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range_obj:
                    if cell.row == merged_range_obj.min_row and cell.column == merged_range_obj.min_col:
                        cell_val_len = len(str(worksheet.cell(row=merged_range_obj.min_row, column=merged_range_obj.min_col).value or ""))
                    is_merged = True
                    break
            if not is_merged and cell.value is not None: cell_val_len = len(str(cell.value or ""))
            if cell_val_len > max_length: max_length = cell_val_len
        worksheet.column_dimensions[column_letter].width = min(max(max_length + padding, 8), max_width)
    debug_print(f"Auto-fitted columns ({', '.join(columns_to_fit) if columns_to_fit else 'All'}) for sheet: {worksheet.title}")

async def async_process_all_ufgs_data(run_label:str = "default_run", concurrency_level_override=None, save_excel=True):
    # Use run-specific variables for counters and maps
    current_run_error_counts = {}
    current_run_global_id_counter = 1
    current_run_df_urls = pd.DataFrame()
    current_run_validation_results_map = {}

    script_start_time = time.time()
    print(f"--- [{run_label}] Main Async Processing Started at {time.ctime(script_start_time)} (Debug: {debug}) ---")
    try:
        import reachable as r_check_main
        print(f"Reachable library version (main script): {getattr(r_check_main, '__version__', '0.7.0 (User Confirmed)')}")
    except ImportError:
        print("Could not import 'reachable' for version check in main script.")
    print(f"httpx library version: {getattr(httpx, '__version__', 'N/A')}")

    if run_label == "1_worker" or (run_label != "100_workers_use_cache" and not os.path.exists(zip_file_path)):
        if not ufgs_download_url or "REPLACE_WITH_CORRECT_UFGS_ZIP_DOWNLOAD_LINK" in ufgs_download_url:
            print(f"[{run_label}] Error: UFGS download URL not set. Exiting.")
            return pd.DataFrame()
        debug_print(f"[{run_label}] Downloading from: {ufgs_download_url}")
        try:
            with httpx.Client(verify=False) as client:
                with client.stream("GET", ufgs_download_url, follow_redirects=True, timeout=60.0) as response:
                    response.raise_for_status()
                    with open(zip_file_path, 'wb') as f:
                        [f.write(chunk) for chunk in response.iter_bytes(8192)]
            print(f"[{run_label}] Download complete: {zip_file_path}")
        except Exception as e:
            print(f"[{run_label}] Download error: {e}")
            return pd.DataFrame()
    elif os.path.exists(zip_file_path):
        print(f"[{run_label}] Using existing/cached zip file: {zip_file_path}")
    else:
        print(f"[{run_label}] Zip file logic error. Exiting.")
        return pd.DataFrame()

    if not os.path.exists(zip_file_path):
        print(f"[{run_label}] Zip file not found: {zip_file_path}")
        return pd.DataFrame()
    if os.path.exists(extract_path):
        shutil.rmtree(extract_path)
    os.makedirs(extract_path)
    debug_print(f"[{run_label}] Extracting zip file...")
    try:
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            zip_ref.extractall(extract_path)
        debug_print(f"[{run_label}] Extraction complete.")
    except Exception as e:
        print(f"[{run_label}] Zip extraction error: {e}")
        return pd.DataFrame()

    all_url_rows_list = []
    sec_files_found_list = [os.path.join(r, i) for r, _, f in os.walk(extract_path) for i in f if i.lower().endswith('.sec')]
    if not sec_files_found_list:
        print(f"[{run_label}] No '.SEC' files found in {extract_path}.")
        return pd.DataFrame()
    print(f"[{run_label}] Found {len(sec_files_found_list)} .SEC files to process.")
    sec_files_found_list.sort()

    file_metadata_list_collector = []
    for file_path in sec_files_found_list:
        debug_print(f"  --- [{run_label}] Parsing file: {os.path.basename(file_path)} ---")
        file_metadata_list_collector.append(extract_file_metadata(file_path))
        url_rows_temp, current_run_global_id_counter = parse_sec_file_for_urls(file_path, current_run_global_id_counter)
        all_url_rows_list.extend(url_rows_temp)

    df_file_metadata = pd.DataFrame(file_metadata_list_collector)
    initial_url_cols = ['ID', 'SEC_FILE_NAME', 'DOC_LINE_NUMBER', 'DOC_CHAR_OFFSET', 'URL_TXT', 'URL_HREF', 'PRIMARY_URL']
    if not all_url_rows_list:
        print(f"[{run_label}] Warning: No URLs found.")
    current_run_df_urls = pd.DataFrame(all_url_rows_list, columns=initial_url_cols) if all_url_rows_list else pd.DataFrame(columns=initial_url_cols)

    df_specs = pd.DataFrame(TARGET_SPECS_DATA, columns=['UFGS', 'Discipline_Spec'])
    current_run_df_urls['Base_Spec'] = current_run_df_urls['SEC_FILE_NAME'].apply(lambda fn: fn[:-4] if isinstance(fn, str) and fn.lower().endswith('.sec') else fn)
    current_run_df_urls = pd.merge(current_run_df_urls, df_specs, left_on='Base_Spec', right_on='UFGS', how='left')
    current_run_df_urls.rename(columns={'Discipline_Spec': 'Discipline'}, inplace=True)
    current_run_df_urls['Discipline'] = current_run_df_urls['Discipline'].fillna('Other')

    def assign_division_01(row):
        sec_file_name_str = str(row.get('SEC_FILE_NAME', ''))
        if re.match(r"^01(\s|[^a-zA-Z0-9_.-]|$)", sec_file_name_str):
            return "DIVISION 01"
        return row['Discipline']

    if 'SEC_FILE_NAME' in current_run_df_urls.columns and 'Discipline' in current_run_df_urls.columns:
        current_run_df_urls['Discipline'] = current_run_df_urls.apply(assign_division_01, axis=1)

    current_run_df_urls.drop(columns=['Base_Spec', 'UFGS'], inplace=True)

    unique_primary_urls_raw = current_run_df_urls['PRIMARY_URL'].dropna().unique().tolist()
    email_urls_for_map_collector = []
    unique_primary_urls_for_batch_collector = []
    invalid_urls_for_map_collector = {}

    for u_url_str_raw in unique_primary_urls_raw:
        u_url_str = str(u_url_str_raw).strip()
        if not u_url_str:
            continue
        if is_email(u_url_str): email_urls_for_map_collector.append(u_url_str)
        else:
            sanitized_check = re.sub(r'[^\x20-\x7E]+', '', u_url_str).strip()
            sanitized_check = re.sub(r'<[^>]+>', '', sanitized_check).strip()
            if not sanitized_check:
                invalid_urls_for_map_collector[u_url_str] = parse_reachable_result({'original_url': u_url_str, 'is_invalid_flag_for_parser': True}, u_url_str, current_run_error_counts)
            else: unique_primary_urls_for_batch_collector.append(u_url_str)

    for email_url in email_urls_for_map_collector:
        current_run_validation_results_map[email_url] = parse_reachable_result({'original_url': email_url, 'is_email_flag_for_parser': True}, email_url, current_run_error_counts)
    for invalid_url, invalid_result_tuple in invalid_urls_for_map_collector.items():
        current_run_validation_results_map[invalid_url] = invalid_result_tuple

    print(f"[{run_label}] Prepared {len(unique_primary_urls_for_batch_collector)} URLs for async batch validation.")
    url_val_start_time = time.time()

    batch_results_list_data = []
    if unique_primary_urls_for_batch_collector:
        CONCURRENCY = concurrency_level_override if concurrency_level_override is not None \
                      else (100 if len(unique_primary_urls_for_batch_collector) > 500 else 50)
        print(f"[{run_label}] Starting async validation (Concurrency: {CONCURRENCY}, using library default timeouts)...")
        batch_results_list_data = await run_async_validators(unique_primary_urls_for_batch_collector, pool_size=CONCURRENCY)
        print(f"[{run_label}] Async URL Validation finished in {time.time() - url_val_start_time:.2f}s.")

        if isinstance(batch_results_list_data, list):
            print(f"[{run_label}] Processing {len(batch_results_list_data)} async results...")
            results_by_url_map = {res.get('original_url'): res for res in batch_results_list_data if isinstance(res, dict) and 'original_url' in res}
            for url_k in unique_primary_urls_for_batch_collector:
                res_d = results_by_url_map.get(url_k)
                if res_d:
                    current_run_validation_results_map[url_k] = parse_reachable_result(res_d, url_k, current_run_error_counts)
                else:
                    current_run_validation_results_map[url_k] = ("BATCH_MISSING_RESULT", None, url_k, "Result missing/unmappable", is_wbdg_url(url_k), "LOW")
                    if url_k not in current_run_error_counts:
                        current_run_error_counts[url_k] = 0
                    current_run_error_counts[url_k] += 1
        else:
            debug_print(f"[{run_label}] Async validation unexpected result type: {type(batch_results_list_data)}.")
            for url_v in unique_primary_urls_for_batch_collector:
                current_run_validation_results_map[url_v] = ("ASYNC_ERROR", None, url_v, f"Async bad type: {type(batch_results_list_data)}", is_wbdg_url(url_v), "LOW")
                if url_v not in current_run_error_counts:
                    current_run_error_counts[url_v] = 0
                current_run_error_counts[url_v] += 1
    else:
        print(f"[{run_label}] No non-email, valid URLs for async batch.")

    result_cols = ['STATUS', 'RESPONSE_CODE', 'FINAL_URL', 'ERROR_MSG', 'IS_WBDG', 'CHECK_CERTAINTY']
    if not current_run_df_urls.empty:
        current_run_df_urls['PRIMARY_URL_STR'] = current_run_df_urls['PRIMARY_URL'].astype(str).str.strip()
        val_series = current_run_df_urls['PRIMARY_URL_STR'].map(current_run_validation_results_map)
        def get_res_part(item, idx):
            if pd.isna(item) or not isinstance(item, tuple) or len(item) != 6:
                col_nm = result_cols[idx] if idx < len(result_cols) else "UNKNOWN_COLUMN"
                if col_nm == 'IS_WBDG':
                    return False
                if col_nm == 'CHECK_CERTAINTY':
                    return "UNKNOWN"
                if col_nm == 'STATUS':
                    return "NOT_MAPPED"
                return None
            return item[idx]
        for i, col_nm in enumerate(result_cols):
            current_run_df_urls[col_nm] = val_series.apply(lambda x: get_res_part(x, i))

        for col, def_val, dtype in [('IS_WBDG', False, bool), ('CHECK_CERTAINTY', "UNKNOWN", str), ('STATUS', "NOT_MAPPED", str)]:
            if col in current_run_df_urls.columns:
                current_run_df_urls[col] = current_run_df_urls[col].fillna(def_val).astype(dtype)
            else:
                current_run_df_urls[col] = def_val

        if 'PRIMARY_URL_STR' in current_run_df_urls.columns and not current_run_df_urls.empty :
            if not current_run_df_urls['PRIMARY_URL_STR'].empty:
                current_run_df_urls['COUNT'] = current_run_df_urls.groupby('PRIMARY_URL_STR')['PRIMARY_URL_STR'].transform('count').fillna(1).astype(int)
            else:
                current_run_df_urls['COUNT'] = 1
            current_run_df_urls.drop(columns=['PRIMARY_URL_STR'], inplace=True, errors='ignore')
        else:
            current_run_df_urls['COUNT'] = 1
    else:
        excel_output_cols_list = initial_url_cols + result_cols + ['Discipline', 'COUNT']
        current_run_df_urls = pd.DataFrame(columns=excel_output_cols_list)

    if save_excel:
        run_specific_excel_path = f"{output_excel_path_base}_{run_label}.xlsx"
        debug_print(f"[{run_label}] Saving DataFrame to Excel: {run_specific_excel_path}")
        try:
            with pd.ExcelWriter(run_specific_excel_path, engine='openpyxl') as writer:
                summary_sheet_name = 'Summary'
                main_sheet_name = 'URL_Checks'
                file_list_sheet_name = 'File_List'
                spec_list_sheet_name = 'Spec_List'
                # New column order for Excel output
                output_columns_main = ['ID', 'SEC_FILE_NAME', 'Discipline', 'STATUS', 'COUNT',
                                       'DOC_LINE_NUMBER', 'DOC_CHAR_OFFSET', 'IS_WBDG',
                                       'CHECK_CERTAINTY', 'PRIMARY_URL', 'RESPONSE_CODE',
                                       'URL_TXT', 'FINAL_URL', 'ERROR_MSG']
                output_columns_filelist = ['Filename', 'Section Number', 'Section Title', 'Date', 'Preparing Authority']
                ws_summary = writer.book.create_sheet(summary_sheet_name, 0)
                header_font = Font(bold=False, size=24)
                sub_header_font = Font(bold=True)
                percent_format_string = '0.0%'
                header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
                grey_fill_summary_label = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center', vertical='center')
                left_align = Alignment(horizontal='left', vertical='center')
                current_row = 1
                overall_header_cell = ws_summary.cell(row=current_row, column=1, value=f"Overall Summary ({run_label})")
                overall_header_cell.font = header_font
                overall_header_cell.fill = header_fill
                ws_summary.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                current_row += 1
                overall_cells = {}
                overall_start_row = current_row

                current_summary_data_for_excel = {}
                if not current_run_df_urls.empty and 'PRIMARY_URL' in current_run_df_urls.columns:
                    total_urls_in_df = len(current_run_df_urls)
                    unique_urls_in_df = current_run_df_urls['PRIMARY_URL'].nunique()
                    fail_status_df = current_run_df_urls[current_run_df_urls['STATUS'] == 'FAIL']
                    total_fail_status = len(fail_status_df)
                    unique_fail_status = fail_status_df['PRIMARY_URL'].nunique()
                    wbdg_content_error_df = current_run_df_urls[current_run_df_urls['STATUS'] == 'WARN_WBDG_CONTENT_ERROR']
                    total_wbdg_content_errors = len(wbdg_content_error_df)
                    unique_wbdg_content_errors = wbdg_content_error_df['PRIMARY_URL'].nunique()
                    total_error_counted_occurrences = sum(current_run_error_counts.values())
                    unique_error_counted_urls = len(current_run_error_counts)
                    current_summary_data_for_excel['Overall'] = {
                        'Total URL Entries Processed (in DataFrame)': total_urls_in_df,
                        'Unique Primary URLs Encountered': unique_urls_in_df,
                        'Total URL Failures (Status: FAIL)': total_fail_status,
                        'Total WBDG Content Warnings (Status: WARN_WBDG_CONTENT_ERROR)': total_wbdg_content_errors,
                        'Total Error Occurrences (from error_counts dict)': total_error_counted_occurrences,
                        'Unique URL Failures (Status: FAIL)': unique_fail_status,
                        'Unique WBDG Content Warnings': unique_wbdg_content_errors,
                        'Unique URLs with Errors (from error_counts dict)': unique_error_counted_urls,
                        '% Total Entries with Failures (Status: FAIL)': None,
                        '% Unique URLs with Failures (Status: FAIL)': None,
                        '% Total Entries with WBDG Warnings': None,
                        '% Unique URLs with WBDG Warnings': None
                    }
                    discipline_summary_excel = {}
                    if 'Discipline' in current_run_df_urls.columns:
                        grouped = current_run_df_urls.groupby('Discipline')
                        for name, group in grouped:
                            disc_total_urls = len(group)
                            disc_unique_urls = group['PRIMARY_URL'].nunique()
                            disc_fail_status_df = group[group['STATUS'] == 'FAIL']
                            disc_total_fail_status = len(disc_fail_status_df)
                            disc_unique_fail_status = disc_fail_status_df['PRIMARY_URL'].nunique()
                            disc_wbdg_warn_df = group[group['STATUS'] == 'WARN_WBDG_CONTENT_ERROR']
                            disc_total_wbdg_warns = len(disc_wbdg_warn_df)
                            disc_unique_wbdg_warns = disc_wbdg_warn_df['PRIMARY_URL'].nunique()
                            discipline_summary_excel[name] = {
                                'Total URLs': disc_total_urls,
                                'Total Failures (FAIL)': disc_total_fail_status,
                                'Total WBDG Content Warnings': disc_total_wbdg_warns,
                                'Unique URLs': disc_unique_urls,
                                'Unique Failures (FAIL)': disc_unique_fail_status,
                                'Unique WBDG Content Warnings': disc_unique_wbdg_warns,
                                '% Total URLs Failure (FAIL)': None,
                                '% Unique URLs Failure (FAIL)': None,
                                '% Total URLs WBDG Warning': None,
                                '% Unique URLs WBDG Warning': None
                            }
                    discipline_order_map = {"DIVISION 01": 0, "Electrical": 1}
                    current_summary_data_for_excel['By Discipline'] = dict(
                        sorted(
                            discipline_summary_excel.items(),
                            key=lambda item: (item[0] == 'Other', discipline_order_map.get(item[0], 2), item[0])
                        )
                    )
                else:
                    current_summary_data_for_excel['Overall'] = {
                        'Total URL Entries Processed (in DataFrame)': 0,
                        'Unique Primary URLs Encountered': 0,
                        'Total URL Failures (Status: FAIL)': 0,
                        'Total WBDG Content Warnings (Status: WARN_WBDG_CONTENT_ERROR)': 0,
                        'Total Error Occurrences (from error_counts dict)': 0,
                        'Unique URL Failures (Status: FAIL)': 0,
                        'Unique WBDG Content Warnings': 0,
                        'Unique URLs with Errors (from error_counts dict)': 0,
                        '% Total Entries with Failures (Status: FAIL)': 0.0,
                        '% Unique URLs with Failures (Status: FAIL)': 0.0,
                        '% Total Entries with WBDG Warnings': 0.0,
                        '% Unique URLs with WBDG Warnings': 0.0
                    }
                    current_summary_data_for_excel['By Discipline'] = {}

                overall_keys_ordered = [
                    'Total URL Entries Processed (in DataFrame)',
                    'Unique Primary URLs Encountered',
                    'Total URL Failures (Status: FAIL)',
                    'Total WBDG Content Warnings (Status: WARN_WBDG_CONTENT_ERROR)',
                    'Total Error Occurrences (from error_counts dict)',
                    'Unique URL Failures (Status: FAIL)',
                    'Unique WBDG Content Warnings',
                    'Unique URLs with Errors (from error_counts dict)',
                    '% Total Entries with Failures (Status: FAIL)',
                    '% Unique URLs with Failures (Status: FAIL)',
                    '% Total Entries with WBDG Warnings',
                    '% Unique URLs with WBDG Warnings'
                ]
                for key in overall_keys_ordered:
                    value = current_summary_data_for_excel['Overall'].get(
                        key,
                        0 if 'URL' in key or 'Failures' in key or 'Warnings' in key or 'Error' in key else (0.0 if '%' in key else "N/A")
                    )
                    label_cell = ws_summary.cell(row=current_row, column=1, value=key)
                    label_cell.alignment = left_align
                    label_cell.fill = grey_fill_summary_label
                    value_cell = ws_summary.cell(row=current_row, column=2, value=value)
                    overall_cells[key] = value_cell.coordinate
                    value_cell.alignment = center_align
                    current_row += 1

                def set_percentage_formula(ws, tc, nc, dc, dv):
                    cell = ws[tc]
                    cell.value = f'=IFERROR({nc}/{dc},0)' if dv > 0 else 0.0
                    cell.number_format = percent_format_string
                    cell.alignment = center_align

                set_percentage_formula(
                    ws_summary,
                    overall_cells['% Total Entries with Failures (Status: FAIL)'],
                    overall_cells["Total URL Failures (Status: FAIL)"],
                    overall_cells["Total URL Entries Processed (in DataFrame)"],
                    current_summary_data_for_excel['Overall'].get('Total URL Entries Processed (in DataFrame)', 0)
                )
                set_percentage_formula(
                    ws_summary,
                    overall_cells['% Unique URLs with Failures (Status: FAIL)'],
                    overall_cells["Unique URL Failures (Status: FAIL)"],
                    overall_cells["Unique Primary URLs Encountered"],
                    current_summary_data_for_excel['Overall'].get('Unique Primary URLs Encountered', 0)
                )
                set_percentage_formula(
                    ws_summary,
                    overall_cells['% Total Entries with WBDG Warnings'],
                    overall_cells["Total WBDG Content Warnings (Status: WARN_WBDG_CONTENT_ERROR)"],
                    overall_cells["Total URL Entries Processed (in DataFrame)"],
                    current_summary_data_for_excel['Overall'].get('Total URL Entries Processed (in DataFrame)', 0)
                )
                set_percentage_formula(
                    ws_summary,
                    overall_cells['% Unique URLs with WBDG Warnings'],
                    overall_cells["Unique WBDG Content Warnings"],
                    overall_cells["Unique Primary URLs Encountered"],
                    current_summary_data_for_excel['Overall'].get('Unique Primary URLs Encountered', 0)
                )
                overall_end_row = current_row - 1
                current_row += 1
                discipline_header_cell = ws_summary.cell(row=current_row, column=1, value="Summary by Discipline")
                discipline_header_cell.font = header_font
                discipline_header_cell.fill = header_fill
                ws_summary.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
                [ws_summary.cell(row=current_row, column=c).fill for c in range(1, 12)]
                current_row += 1
                discipline_start_row = current_row
                if current_summary_data_for_excel['By Discipline']:
                    disc_headers_ordered = [
                        'Total URLs',
                        'Total Failures (FAIL)',
                        'Total WBDG Content Warnings',
                        'Unique URLs',
                        'Unique Failures (FAIL)',
                        'Unique WBDG Content Warnings',
                        '% Total URLs Failure (FAIL)',
                        '% Unique URLs Failure (FAIL)',
                        '% Total URLs WBDG Warning',
                        '% Unique URLs WBDG Warning'
                    ]
                    ws_summary.cell(row=current_row, column=1, value="Discipline").font = sub_header_font
                    ws_summary.cell(row=current_row, column=1).fill = grey_fill_summary_label
                    ws_summary.cell(row=current_row, column=1).border = thin_border
                    ws_summary.cell(row=current_row, column=1).alignment = left_align
                    for col_idx, header in enumerate(disc_headers_ordered):
                        cell = ws_summary.cell(row=current_row, column=col_idx + 2, value=header)
                        cell.font = sub_header_font
                        cell.fill = grey_fill_summary_label
                        cell.border = thin_border
                        cell.alignment = center_align
                    current_row += 1
                    for discipline, stats in current_summary_data_for_excel['By Discipline'].items():
                        ws_summary.cell(row=current_row, column=1, value=discipline).font = sub_header_font
                        ws_summary.cell(row=current_row, column=1).border = thin_border
                        ws_summary.cell(row=current_row, column=1).alignment = left_align
                        discipline_cells = {}
                        col_idx = 2
                        for key in disc_headers_ordered:
                            value = stats.get(
                                key,
                                0 if 'URL' in key or 'Failures' in key or 'Warning' in key else (0.0 if '%' in key else "N/A")
                            )
                            cell = ws_summary.cell(row=current_row, column=col_idx, value=value)
                            discipline_cells[key] = cell.coordinate
                            cell.border = thin_border
                            cell.alignment = center_align
                            col_idx += 1
                        set_percentage_formula(
                            ws_summary,
                            discipline_cells['% Total URLs Failure (FAIL)'],
                            discipline_cells["Total Failures (FAIL)"],
                            discipline_cells["Total URLs"],
                            stats.get('Total URLs', 0)
                        )
                        set_percentage_formula(
                            ws_summary,
                            discipline_cells['% Unique URLs Failure (FAIL)'],
                            discipline_cells["Unique Failures (FAIL)"],
                            discipline_cells["Unique URLs"],
                            stats.get('Unique URLs', 0)
                        )
                        set_percentage_formula(
                            ws_summary,
                            discipline_cells['% Total URLs WBDG Warning'],
                            discipline_cells["Total WBDG Content Warnings"],
                            discipline_cells["Total URLs"],
                            stats.get('Total URLs', 0)
                        )
                        set_percentage_formula(
                            ws_summary,
                            discipline_cells['% Unique URLs WBDG Warning'],
                            discipline_cells["Unique WBDG Content Warnings"],
                            discipline_cells["Unique URLs"],
                            stats.get('Unique URLs', 0)
                        )
                        current_row += 1
                else:
                    ws_summary.cell(row=current_row, column=1, value="No discipline data found.")
                    current_row += 1
                discipline_end_row = current_row - 1
                [[ws_summary.cell(row=r, column=c).border for c in range(1, 3)] for r in range(overall_start_row, overall_end_row + 1)]
                [[ws_summary.cell(row=r_idx, column=c_idx_d).border for c_idx_d in range(1, 12)] for r_idx in range(discipline_start_row, discipline_end_row + 1) if current_summary_data_for_excel['By Discipline']]

                if current_run_df_urls.empty: df_to_save = pd.DataFrame(columns=output_columns_main)
                else:
                    for col_main in output_columns_main:
                        if col_main not in current_run_df_urls.columns: current_run_df_urls[col_main] = None
                    df_to_save = current_run_df_urls[output_columns_main].copy()

                df_to_save = df_to_save.fillna('')
                df_to_save.to_excel(writer, index=False, sheet_name=main_sheet_name)
                ws_main = writer.sheets[main_sheet_name]
                ws_main.freeze_panes = "A2"
                df_file_metadata.to_excel(writer, index=False, sheet_name=file_list_sheet_name)
                ws_file_list = writer.sheets[file_list_sheet_name]
                ws_file_list.freeze_panes = "A2"
                if not df_file_metadata.empty:
                    table_files = Table(displayName="FileListTable", ref=f"A1:{get_column_letter(len(output_columns_filelist))}{len(df_file_metadata)+1}")
                    table_files.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    _ = [ws_file_list.add_table(table_files) for t in ws_file_list._tables if t.name != table_files.displayName] if not any(t.name == table_files.displayName for t in ws_file_list._tables) else None
                ws_spec_list = writer.book.create_sheet(spec_list_sheet_name)
                ws_spec_list['A1'] = "UFGS"
                ws_spec_list['B1'] = "Discipline"
                # Corrected line for writing spec list data
                for i, (spec, discipline_val) in enumerate(TARGET_SPECS_DATA):
                    ws_spec_list.cell(row=i + 2, column=1, value=spec)
                    ws_spec_list.cell(row=i + 2, column=2, value=discipline_val)
                ws_spec_list.freeze_panes = "A2"
                if TARGET_SPECS_DATA:
                    table_spec = Table(displayName="SpecListTable", ref=f"A1:B{len(TARGET_SPECS_DATA)+1}")
                    table_spec.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    _ = [ws_spec_list.add_table(table_spec) for t in ws_spec_list._tables if t.name != table_spec.displayName] if not any(t.name == table_spec.displayName for t in ws_spec_list._tables) else None

                red_fill = PatternFill(start_color='FFC7CE', fill_type='solid')
                red_font = Font(color='9C0006')
                green_fill = PatternFill(start_color='C6EFCE', fill_type='solid')
                green_font = Font(color='006100')
                orange_fill = PatternFill(start_color='FFD9B3', fill_type='solid')
                orange_font = Font(color='A65B00')
                grey_fill_error = PatternFill(start_color='D9D9D9', fill_type='solid')
                grey_font_error = Font(color='595959')

                if len(df_to_save) > 0:
                    formatting_range_main = f"A2:{get_column_letter(len(output_columns_main))}{len(df_to_save) + 1}"
                    try:
                        status_col_letter_main = get_column_letter(output_columns_main.index('STATUS') + 1)
                        for f_str, font, fill in [(f'${status_col_letter_main}2="PASS"', green_font, green_fill), (f'${status_col_letter_main}2="FAIL"', red_font, red_fill), (f'${status_col_letter_main}2="WARN_WBDG_CONTENT_ERROR"', orange_font, orange_fill), (f'OR(${status_col_letter_main}2="PROCESSING_ERROR", ${status_col_letter_main}2="NOT_MAPPED", ${status_col_letter_main}2="BATCH_ITEM_ERROR", ${status_col_letter_main}2="BATCH_LENGTH_ERROR", ${status_col_letter_main}2="BATCH_TYPE_ERROR", ${status_col_letter_main}2="BATCH_CALL_EXCEPTION", ${status_col_letter_main}2="BATCH_MISSING_RESULT", ${status_col_letter_main}2="ASYNC_OVERALL_ERROR", ${status_col_letter_main}2="ERROR_IN_LOOP", ${status_col_letter_main}2="FAIL_WORKER")', grey_font_error, grey_fill_error)]: ws_main.conditional_formatting.add(formatting_range_main, FormulaRule(formula=[f_str], stopIfTrue=False, font=font, fill=fill))
                    except ValueError:
                        debug_print("Error finding 'STATUS' column for conditional formatting.")
                if output_columns_main and len(df_to_save) > 0:
                    table_main = Table(displayName="URLCheckTable", ref=f"A1:{get_column_letter(len(output_columns_main))}{len(df_to_save) + 1}")
                    table_main.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    _ = [ws_main.add_table(table_main) for t in ws_main._tables if t.name != table_main.displayName] if not any(t.name == table_main.displayName for t in ws_main._tables) else None
                auto_fit_columns(ws_summary, columns_to_fit=['A', 'B'])
                auto_fit_columns(ws_main)
                auto_fit_columns(ws_file_list)
                auto_fit_columns(ws_spec_list)
                desired_order = [summary_sheet_name, main_sheet_name, file_list_sheet_name, spec_list_sheet_name]
                current_sheet_titles = [s.title for s in writer.book._sheets]
                writer.book._sheets = (
                    sorted(
                        writer.book._sheets,
                        key=lambda s: (desired_order.index(s.title) if s.title in desired_order else len(desired_order))
                    )
                    if all(sn in current_sheet_titles for sn in desired_order)
                    else writer.book._sheets
                )
                debug_print(f"[{run_label}] Excel file created: {run_specific_excel_path}")
        except Exception as e_excel:
            print(f"[{run_label}] Excel saving error: {e_excel}")
            traceback.print_exc()

    final_proc_time = time.time() - script_start_time
    print(f"\n[{run_label}] Total Processing Time: {time.strftime('%H:%M:%S', time.gmtime(final_proc_time))}")
    print(f"\n[{run_label}] URL Validation Issue Summary (from error_counts dict for this run):")
    if current_run_error_counts:
        sorted_errors = sorted(current_run_error_counts.items(), key=lambda item: (-item[1], item[0]))
        for url_key, count_val in sorted_errors:
            err_tuple = current_run_validation_results_map.get(url_key)
            if err_tuple and isinstance(err_tuple, tuple) and len(err_tuple) == 6:
                 status_v, code_v, _, msg_v, _, cert_v = err_tuple
                 print(f"  - Cnt: {count_val}, Status: {status_v}, Code: {code_v}, URL: {url_key[:70]}..., Msg: {msg_v}, Cert: {cert_v}")
            else:
                 status_val = err_tuple[0] if err_tuple and isinstance(err_tuple, tuple) and len(err_tuple) > 0 else "N/A"
                 msg_val = err_tuple[3] if err_tuple and isinstance(err_tuple, tuple) and len(err_tuple) > 3 else "N/A"
                 print(f"  - Cnt: {count_val}, Status: {status_val}, URL: {url_key[:70]}... (Detail limited: {msg_val})")
    else:
        total_urls_for_validation = len(unique_primary_urls_for_batch_collector) + len(email_urls_for_map_collector) + len(invalid_urls_for_map_collector)
        print(f"  [{run_label}] No URL validation issues recorded in error_counts for this run." if total_urls_for_validation > 0 else f"  [{run_label}] URL validation skipped (no valid URLs found).")
    print(f"\n[{run_label}] Analysis complete. Output (if saved): {run_specific_excel_path if save_excel else 'Excel saving disabled for this run.'}")

    return current_run_df_urls

async def run_concurrency_comparison_test():
    if not async_components_imported_successfully:
        print("CRITICAL: Cannot run comparison test as async components from 'reachable' failed to import.")
        return

    print("\n--- Starting Concurrency Comparison Test ---")

    if not os.path.exists(zip_file_path):
        print("Initial zip download for comparison test...")
        if not ufgs_download_url or "REPLACE_WITH_CORRECT_UFGS_ZIP_DOWNLOAD_LINK" in ufgs_download_url:
            print("Error: UFGS download URL not set. Exiting test.")
            return
        try:
            with httpx.Client(verify=False) as client:
                with client.stream("GET", ufgs_download_url, follow_redirects=True, timeout=60.0) as response:
                    response.raise_for_status()
                    with open(zip_file_path, 'wb') as f: [f.write(chunk) for chunk in response.iter_bytes(8192)]
            print(f"Initial download complete: {zip_file_path}")
        except Exception as e:
            print(f"Initial download error: {e}")
            return

    print("\n--- Running with 1 Worker ---")
    df_run1 = await async_process_all_ufgs_data(run_label="1_worker", concurrency_level_override=1, save_excel=False)

    print("\n--- Running with 100 Workers ---")
    df_run2 = await async_process_all_ufgs_data(run_label="100_workers_use_cache", concurrency_level_override=100, save_excel=False)

    print("\n\n--- DataFrame Comparison Results ---")
    if df_run1.empty and df_run2.empty:
        print("Both runs resulted in empty DataFrames. Nothing to compare.")
        return
    if df_run1.empty:
        print("DataFrame from 1-worker run is empty. Cannot compare.")
        return
    if df_run2.empty:
        print("DataFrame from 100-worker run is empty. Cannot compare.")
        return
    if 'PRIMARY_URL' not in df_run1.columns or 'PRIMARY_URL' not in df_run2.columns:
        print("Error: 'PRIMARY_URL' column missing. Cannot compare.")
        return

    compare_cols = ['STATUS', 'RESPONSE_CODE', 'FINAL_URL', 'ERROR_MSG', 'CHECK_CERTAINTY', 'IS_WBDG']
    for df_temp in [df_run1, df_run2]:
        for col_c in compare_cols:
            if col_c not in df_temp.columns:
                df_temp[col_c] = pd.NA

    merged_df = pd.merge(
        df_run1[['PRIMARY_URL'] + compare_cols].fillna("DF1_NA_FILL"),
        df_run2[['PRIMARY_URL'] + compare_cols].fillna("DF2_NA_FILL"),
        on='PRIMARY_URL', how='outer', suffixes=('_1w', '_100w')
    )

    differences_found = 0
    print(f"Comparing {len(merged_df)} unique URLs found across both runs...")
    for _, row in merged_df.iterrows():
        url = row['PRIMARY_URL']
        discrepancy_details = []
        for col_base in compare_cols:
            val1, val100 = row[f"{col_base}_1w"], row[f"{col_base}_100w"]
            s_val1 = str(val1 if not pd.isna(val1) else "MISSING_IN_RUN1")
            s_val100 = str(val100 if not pd.isna(val100) else "MISSING_IN_RUN2")
            if s_val1 != s_val100:
                discrepancy_details.append(f"  - {col_base}: 1w='{s_val1}' vs 100w='{s_val100}'")
        if discrepancy_details:
            differences_found += 1
            print(f"\nDiscrepancy for URL: {url}")
            for detail in discrepancy_details:
                print(detail)

    if differences_found == 0:
        print("\nNo differences found in validation results between 1-worker and 100-worker runs.")
    else:
        print(f"\nFound differences for {differences_found} URLs.")
    print("\n--- Concurrency Comparison Test Finished ---")

# --- Wrapper function to be called from a new Colab cell using await ---
async def start_processing_if_configured(concurrency_override=None, run_comparison_test=False):
    # Ensure global DO_FULL_PROCESSING is accessible if not passed
    global DO_FULL_PROCESSING, async_components_imported_successfully

    if DO_FULL_PROCESSING and async_components_imported_successfully:
        if run_comparison_test:
            await run_concurrency_comparison_test()
        else:
            await async_process_all_ufgs_data(
                run_label=f"run_concurrency_{concurrency_override if concurrency_override else 'default'}",
                concurrency_level_override=concurrency_override,
                save_excel=True
            )
    else:
        if not DO_FULL_PROCESSING:
            print("DO_FULL_PROCESSING is False. Set to True in the script and re-run this cell if you want to process data.")
        if not async_components_imported_successfully:
            print("CRITICAL: Async components from 'reachable' were not imported successfully. The main process cannot run.")

print("\nScript definitions complete. In a new cell, run: await start_processing_if_configured()")
print("Or for comparison test: await start_processing_if_configured(run_comparison_test=True)")

# In[2]:
if DO_FULL_PROCESSING and async_components_imported_successfully:
    # This will run the main process with default concurrency
    # (50 or 100 based on URL count) and save the Excel output.
    await async_process_all_ufgs_data(run_label="final_output", save_excel=True)

    # --- Optional: For testing specific concurrency levels ---
    # To test with 1 worker (slower, but good for isolating issues):
    # print("\n--- RUNNING WITH 1 WORKER FOR TESTING ---")
    # await async_process_all_ufgs_data(run_label="1_worker_test", concurrency_level_override=1, save_excel=True)

    # To test with 100 workers (faster, more load):
    # print("\n--- RUNNING WITH 100 WORKERS FOR TESTING ---")
    # await async_process_all_ufgs_data(run_label="100_workers_test", concurrency_level_override=100, save_excel=True)

else:
    if not DO_FULL_PROCESSING:
        print("DO_FULL_PROCESSING is False. Please set it to True in the main script cell and re-run that cell if you want to process data.")
    if not async_components_imported_successfully:
        print("CRITICAL: Async components from 'reachable' were not imported successfully (check output of the first cell). The main process cannot run.")
